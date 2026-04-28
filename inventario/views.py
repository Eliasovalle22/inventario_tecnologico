from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from .models import Activo
from .forms import ActivoForm, ImportarActivosForm
from movimientos.models import Movimiento
from catalogos.models import Estado, TipoActivo, Categoria, Marca, Ubicacion
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl
from datetime import datetime

@login_required
@permission_required('inventario.view_activo', raise_exception=True)
def lista_activos(request):
    # Obtener todos los activos con relaciones
    activos_list = Activo.objects.select_related(
        'tipo', 'categoria', 'marca', 'estado', 'ubicacion', 'responsable'
    ).all().order_by('-fecha_creacion')
    
    # Búsqueda
    query = request.GET.get('q')
    if query:
        activos_list = activos_list.filter(
            Q(codigo__icontains=query) |
            Q(modelo__icontains=query) |
            Q(marca__nombre__icontains=query) |
            Q(responsable__first_name__icontains=query) |
            Q(responsable__last_name__icontains=query)
        )
    
    # Filtros adicionales
    estado = request.GET.get('estado')
    if estado:
        activos_list = activos_list.filter(estado_id=estado)
    
    tipo = request.GET.get('tipo')
    if tipo:
        activos_list = activos_list.filter(tipo_id=tipo)
    
    # Obtener listas para filtros
    estados = Estado.objects.all()
    tipos = TipoActivo.objects.all()
    
    context = {
        'activos': activos_list,
        'estados': estados,
        'tipos': tipos,
        'query': query,
        'filtro_estado': estado,
        'filtro_tipo': tipo,
    }
    return render(request, 'inventario/lista_activos.html', context)

@login_required
@permission_required('inventario.view_activo', raise_exception=True)
def detalle_activo(request, pk):
    activo = get_object_or_404(
        Activo.objects.select_related(
            'tipo', 'categoria', 'marca', 'estado', 'ubicacion', 'responsable', 'creado_por'
        ),
        pk=pk
    )
    
    # Obtener movimientos del activo
    movimientos = activo.movimientos.select_related('usuario').all().order_by('-fecha')[:10]
    
    # Obtener historial de asignaciones
    asignaciones = activo.asignaciones.select_related('usuario_asignado', 'asignado_por').all().order_by('-fecha_asignacion')[:5]
    
    context = {
        'activo': activo,
        'movimientos': movimientos,
        'asignaciones': asignaciones,
    }
    return render(request, 'inventario/detalle_activo.html', context)

@login_required
@permission_required('inventario.add_activo', raise_exception=True)
def crear_activo(request):
    if request.method == 'POST':
        form = ActivoForm(request.POST)
        if form.is_valid():
            activo = form.save(commit=False)
            activo.creado_por = request.user
            activo.save()
            
            # Registrar movimiento de creación
            Movimiento.objects.create(
                activo=activo,
                tipo='ENTRADA',
                usuario=request.user,
                ubicacion_destino=activo.ubicacion,
                estado_destino=activo.estado,
                observaciones=f"Activo creado por {request.user.get_full_name()}"
            )
            
            messages.success(request, f'Activo {activo.codigo} creado exitosamente.')
            
            if 'guardar_y_nuevo' in request.POST:
                return redirect('inventario:crear')
            
            return redirect('inventario:detalle', pk=activo.pk)
    else:
        form = ActivoForm()
    
    context = {
        'form': form,
        'titulo': 'Nuevo Activo',
        'boton': 'Crear Activo'
    }
    return render(request, 'inventario/activo_form.html', context)

@login_required
@permission_required('inventario.change_activo', raise_exception=True)
def editar_activo(request, pk):
    activo = get_object_or_404(Activo, pk=pk)
    old_values = {
        'estado': activo.estado,
        'ubicacion': activo.ubicacion,
        'responsable': activo.responsable,
    }
    
    if request.method == 'POST':
        form = ActivoForm(request.POST, instance=activo)
        if form.is_valid():
            activo_actualizado = form.save()
            
            # Verificar cambios y registrar movimientos
            new_values = {
                'estado': activo_actualizado.estado,
                'ubicacion': activo_actualizado.ubicacion,
                'responsable': activo_actualizado.responsable,
            }
            
            # Registrar cambios significativos
            if old_values['estado'] != new_values['estado']:
                Movimiento.objects.create(
                    activo=activo_actualizado,
                    tipo='CAMBIO_ESTADO',
                    usuario=request.user,
                    estado_origen=old_values['estado'],
                    estado_destino=new_values['estado'],
                    observaciones=f"Cambio de estado por {request.user.get_full_name()}"
                )
            
            if old_values['ubicacion'] != new_values['ubicacion']:
                Movimiento.objects.create(
                    activo=activo_actualizado,
                    tipo='TRASLADO',
                    usuario=request.user,
                    ubicacion_origen=old_values['ubicacion'],
                    ubicacion_destino=new_values['ubicacion'],
                    observaciones=f"Traslado realizado por {request.user.get_full_name()}"
                )
            
            messages.success(request, f'Activo {activo.codigo} actualizado exitosamente.')
            return redirect('inventario:detalle', pk=activo.pk)
    else:
        form = ActivoForm(instance=activo)
    
    context = {
        'form': form,
        'activo': activo,
        'titulo': f'Editar Activo: {activo.codigo}',
        'boton': 'Actualizar Activo'
    }
    return render(request, 'inventario/activo_form.html', context)

@login_required
@permission_required('inventario.delete_activo', raise_exception=True)
def eliminar_activo(request, pk):
    activo = get_object_or_404(Activo, pk=pk)
    
    if request.method == 'POST':
        codigo = activo.codigo
        # Registrar movimiento de baja
        Movimiento.objects.create(
            activo=activo,
            tipo='BAJA',
            usuario=request.user,
            observaciones=f"Activo dado de baja por {request.user.get_full_name()}"
        )
        activo.delete()
        messages.success(request, f'Activo {codigo} eliminado exitosamente.')
        return redirect('inventario:lista')
    
    context = {
        'activo': activo
    }
    return render(request, 'inventario/confirmar_eliminar.html', context)


@login_required
@permission_required('inventario.add_activo', raise_exception=True)
def descargar_plantilla_activos(request):
    """
    Descarga una plantilla de Excel con validación de datos.
    """
    # Crear un nuevo libro
    wb = Workbook()
    ws = wb.active
    ws.title = "Activos"
    
    # Definir estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definir encabezados
    headers = [
        'Código/Inventario',
        'Tipo Activo',
        'Categoría',
        'Marca',
        'Modelo',
        'Estado',
        'Ubicación',
        'Fecha Compra (YYYY-MM-DD)',
        'Valor Compra',
        'Garantía (meses)',
        'Proveedor',
        'N° Factura',
        'Especificaciones',
        'Observaciones'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Establecer ancho de columnas
    column_widths = [18, 18, 18, 18, 20, 15, 15, 20, 15, 15, 20, 15, 30, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # ===== CREAR HOJAS DE REFERENCIA =====
    
    # Tipos de Activos
    tipos = TipoActivo.objects.all().values_list('nombre', flat=True)
    ws_tipos = wb.create_sheet('Tipos de Activo')
    for idx, tipo in enumerate(tipos, 1):
        ws_tipos[f'A{idx}'].value = tipo
    
    # Categorías
    categorias = Categoria.objects.all().values_list('nombre', flat=True)
    ws_categorias = wb.create_sheet('Categorías')
    for idx, categoria in enumerate(categorias, 1):
        ws_categorias[f'A{idx}'].value = categoria
    
    # Marcas
    marcas = Marca.objects.all().values_list('nombre', flat=True)
    ws_marcas = wb.create_sheet('Marcas')
    for idx, marca in enumerate(marcas, 1):
        ws_marcas[f'A{idx}'].value = marca
    
    # Estados
    estados = Estado.objects.all().values_list('nombre', flat=True)
    ws_estados = wb.create_sheet('Estados')
    for idx, estado in enumerate(estados, 1):
        ws_estados[f'A{idx}'].value = estado
    
    # Ubicaciones
    ubicaciones = Ubicacion.objects.all().values_list('nombre', flat=True)
    ws_ubicaciones = wb.create_sheet('Ubicaciones')
    for idx, ubicacion in enumerate(ubicaciones, 1):
        ws_ubicaciones[f'A{idx}'].value = ubicacion
    
    # ===== AGREGAR VALIDACIÓN DE DATOS =====
    
    # Volver a la hoja principal
    ws = wb['Activos']
    
    # Validación para Tipo Activo (columna B, filas 2-1000)
    dv_tipo = DataValidation(
        type="list",
        formula1=f"'Tipos de Activo'!$A$1:$A${len(tipos)}",
        allow_blank=False
    )
    dv_tipo.error = 'Selecciona un tipo válido'
    dv_tipo.errorTitle = 'Tipo inválido'
    ws.add_data_validation(dv_tipo)
    dv_tipo.add(f'B2:B1000')
    
    # Validación para Categoría (columna C, filas 2-1000)
    dv_categoria = DataValidation(
        type="list",
        formula1=f"'Categorías'!$A$1:$A${len(categorias)}",
        allow_blank=False
    )
    dv_categoria.error = 'Selecciona una categoría válida'
    dv_categoria.errorTitle = 'Categoría inválida'
    ws.add_data_validation(dv_categoria)
    dv_categoria.add(f'C2:C1000')
    
    # Validación para Marca (columna D, filas 2-1000)
    dv_marca = DataValidation(
        type="list",
        formula1=f"'Marcas'!$A$1:$A${len(marcas)}",
        allow_blank=False
    )
    dv_marca.error = 'Selecciona una marca válida'
    dv_marca.errorTitle = 'Marca inválida'
    ws.add_data_validation(dv_marca)
    dv_marca.add(f'D2:D1000')
    
    # Validación para Estado (columna F, filas 2-1000)
    dv_estado = DataValidation(
        type="list",
        formula1=f"'Estados'!$A$1:$A${len(estados)}",
        allow_blank=False
    )
    dv_estado.error = 'Selecciona un estado válido'
    dv_estado.errorTitle = 'Estado inválido'
    ws.add_data_validation(dv_estado)
    dv_estado.add(f'F2:F1000')
    
    # Validación para Ubicación (columna G, filas 2-1000)
    dv_ubicacion = DataValidation(
        type="list",
        formula1=f"'Ubicaciones'!$A$1:$A${len(ubicaciones)}",
        allow_blank=False
    )
    dv_ubicacion.error = 'Selecciona una ubicación válida'
    dv_ubicacion.errorTitle = 'Ubicación inválida'
    ws.add_data_validation(dv_ubicacion)
    dv_ubicacion.add(f'G2:G1000')
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Plantilla_Activos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@permission_required('inventario.add_activo', raise_exception=True)
@require_http_methods(["POST"])
def importar_activos(request):
    """
    Importa activos desde un archivo Excel.
    """
    form = ImportarActivosForm(request.POST, request.FILES)
    
    if form.is_valid():
        archivo = request.FILES['archivo']
        
        try:
            # Cargar el archivo Excel
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            # Obtener datos de la hoja
            errores = []
            advertencias = []
            activos_creados = 0
            
            # Procesar filas (comenzar desde la fila 2, la 1 es encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                try:
                    # Obtener valores de las celdas
                    codigo = row[0].value
                    tipo_nombre = row[1].value
                    categoria_nombre = row[2].value
                    marca_nombre = row[3].value
                    modelo = row[4].value
                    estado_nombre = row[5].value
                    ubicacion_nombre = row[6].value
                    fecha_compra = row[7].value
                    valor_compra = row[8].value
                    garantia_meses = row[9].value
                    proveedor = row[10].value
                    factura = row[11].value
                    especificaciones = row[12].value
                    observaciones = row[13].value
                    
                    # Validar que no esté vacío
                    if not codigo:
                        continue
                    
                    # Validar datos requeridos
                    if not all([tipo_nombre, categoria_nombre, marca_nombre, modelo, estado_nombre, ubicacion_nombre]):
                        errores.append(f"Fila {row_num}: Faltan datos requeridos (Tipo, Categoría, Marca, Modelo, Estado, Ubicación)")
                        continue
                    
                    # Obtener objetos relacionados
                    try:
                        tipo = TipoActivo.objects.get(nombre__iexact=str(tipo_nombre).strip())
                    except TipoActivo.DoesNotExist:
                        errores.append(f"Fila {row_num}: Tipo de activo '{tipo_nombre}' no existe")
                        continue
                    
                    try:
                        categoria = Categoria.objects.get(nombre__iexact=str(categoria_nombre).strip())
                    except Categoria.DoesNotExist:
                        errores.append(f"Fila {row_num}: Categoría '{categoria_nombre}' no existe")
                        continue
                    
                    try:
                        marca = Marca.objects.get(nombre__iexact=str(marca_nombre).strip())
                    except Marca.DoesNotExist:
                        errores.append(f"Fila {row_num}: Marca '{marca_nombre}' no existe")
                        continue
                    
                    try:
                        estado = Estado.objects.get(nombre__iexact=str(estado_nombre).strip())
                    except Estado.DoesNotExist:
                        errores.append(f"Fila {row_num}: Estado '{estado_nombre}' no existe")
                        continue
                    
                    try:
                        ubicacion = Ubicacion.objects.get(nombre__iexact=str(ubicacion_nombre).strip())
                    except Ubicacion.DoesNotExist:
                        errores.append(f"Fila {row_num}: Ubicación '{ubicacion_nombre}' no existe")
                        continue
                    
                    # Validar que el código sea único
                    if Activo.objects.filter(codigo__iexact=str(codigo).strip()).exists():
                        errores.append(f"Fila {row_num}: El código '{codigo}' ya existe")
                        continue
                    
                    # Procesar fecha_compra
                    if fecha_compra:
                        try:
                            if isinstance(fecha_compra, datetime):
                                fecha_compra = fecha_compra.date()
                            else:
                                fecha_compra = datetime.strptime(str(fecha_compra), '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            advertencias.append(f"Fila {row_num}: Fecha de compra inválida, se ignorará")
                            fecha_compra = None
                    
                    # Procesar valor_compra
                    if valor_compra:
                        try:
                            valor_compra = float(valor_compra)
                        except (ValueError, TypeError):
                            advertencias.append(f"Fila {row_num}: Valor de compra inválido, se ignorará")
                            valor_compra = None
                    
                    # Procesar garantia_meses
                    if garantia_meses:
                        try:
                            garantia_meses = int(garantia_meses)
                        except (ValueError, TypeError):
                            garantia_meses = 0
                    else:
                        garantia_meses = 0
                    
                    # Crear el activo
                    activo = Activo(
                        codigo=str(codigo).strip().upper(),
                        tipo=tipo,
                        categoria=categoria,
                        marca=marca,
                        modelo=str(modelo).strip() if modelo else '',
                        estado=estado,
                        ubicacion=ubicacion,
                        fecha_compra=fecha_compra,
                        valor_compra=valor_compra,
                        garantia_meses=garantia_meses or 0,
                        proveedor=str(proveedor).strip() if proveedor else None,
                        factura=str(factura).strip() if factura else None,
                        especificaciones=str(especificaciones).strip() if especificaciones else None,
                        observaciones=str(observaciones).strip() if observaciones else None,
                        creado_por=request.user
                    )
                    activo.save()
                    
                    # Registrar movimiento de creación
                    Movimiento.objects.create(
                        activo=activo,
                        tipo='ENTRADA',
                        usuario=request.user,
                        ubicacion_destino=activo.ubicacion,
                        estado_destino=activo.estado,
                        observaciones=f"Activo importado por {request.user.get_full_name()}"
                    )
                    
                    activos_creados += 1
                
                except Exception as e:
                    errores.append(f"Fila {row_num}: Error al procesar - {str(e)}")
            
            # Preparar respuesta
            respuesta = {
                'success': True,
                'activos_creados': activos_creados,
                'errores': errores,
                'advertencias': advertencias,
                'total_errores': len(errores),
                'total_advertencias': len(advertencias)
            }
            
            # Mensaje de resumen
            if activos_creados > 0:
                messages.success(request, f'✓ Se importaron {activos_creados} activo(s) exitosamente.')
            
            if errores:
                mensaje_errores = '\n'.join(errores[:5])
                if len(errores) > 5:
                    mensaje_errores += f'\n... y {len(errores) - 5} error(es) más'
                messages.error(request, f'⚠ Errores encontrados durante la importación:\n{mensaje_errores}')
            
            if advertencias:
                mensaje_advertencias = '\n'.join(advertencias[:3])
                if len(advertencias) > 3:
                    mensaje_advertencias += f'\n... y {len(advertencias) - 3} advertencia(s) más'
                messages.warning(request, f'ℹ {mensaje_advertencias}')
            
            return JsonResponse(respuesta)
        
        except openpyxl.utils.exceptions.InvalidFileException:
            return JsonResponse({
                'success': False,
                'mensaje': 'El archivo no es un archivo Excel válido.'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'mensaje': f'Error al procesar el archivo: {str(e)}'
            }, status=400)
    
    else:
        errores_formulario = form.errors.as_json()
        return JsonResponse({
            'success': False,
            'mensaje': 'Archivo inválido',
            'errores': errores_formulario
        }, status=400)