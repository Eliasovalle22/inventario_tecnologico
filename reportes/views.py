import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import timedelta
from io import BytesIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context
import matplotlib.pyplot as plt
import base64
from matplotlib import rcParams
rcParams['font.family'] = 'sans-serif'
rcParams['font.sans-serif'] = ['Arial']

from inventario.models import Activo
from asignaciones.models import Asignacion
from movimientos.models import Movimiento
from catalogos.models import Categoria, Estado, Ubicacion
from django.contrib.auth.models import User

# ============= DASHBOARD PRINCIPAL DE REPORTES =============
@login_required
@permission_required('reportes.view_reporte', raise_exception=True)
def dashboard_reportes(request):
    """Dashboard principal con gráficos y estadísticas"""
    
    # Estadísticas generales
    total_activos = Activo.objects.count()
    activos_asignados = Activo.objects.filter(estado__nombre='ASIGNADO').count()
    activos_disponibles = Activo.objects.filter(estado__nombre='DISPONIBLE').count()
    activos_baja = Activo.objects.filter(estado__nombre='DADO DE BAJA').count()
    
    # Asignaciones activas
    asignaciones_activas = Asignacion.objects.filter(activo_actual=True).count()
    asignaciones_vencidas = Asignacion.objects.filter(
        activo_actual=True,
        fecha_estimada_devolucion__lt=timezone.now()
    ).count()
    
    # Movimientos últimos 30 días
    fecha_limite = timezone.now() - timedelta(days=30)
    movimientos_30d = Movimiento.objects.filter(fecha__gte=fecha_limite).count()
    
    # Activos por ubicación
    activos_por_ubicacion = []
    for ubicacion in Ubicacion.objects.all()[:5]:  # Top 5
        count = Activo.objects.filter(ubicacion=ubicacion).count()
        if count > 0:
            activos_por_ubicacion.append({
                'nombre': ubicacion.nombre,
                'total': count
            })
    
    context = {
        'total_activos': total_activos,
        'activos_asignados': activos_asignados,
        'activos_disponibles': activos_disponibles,
        'activos_baja': activos_baja,
        'asignaciones_activas': asignaciones_activas,
        'asignaciones_vencidas': asignaciones_vencidas,
        'movimientos_30d': movimientos_30d,
        'activos_por_ubicacion': activos_por_ubicacion,
    }
    return render(request, 'reportes/dashboard.html', context)

# ============= REPORTE DE INVENTARIO =============
@login_required
def reporte_inventario(request):
    """Vista de reporte de inventario con filtros"""
    
    activos = Activo.objects.select_related(
        'tipo', 'categoria', 'marca', 'estado', 'ubicacion', 'responsable'
    ).all()
    
    # Filtros
    estado = request.GET.get('estado')
    categoria = request.GET.get('categoria')
    ubicacion = request.GET.get('ubicacion')
    
    if estado:
        activos = activos.filter(estado_id=estado)
    if categoria:
        activos = activos.filter(categoria_id=categoria)
    if ubicacion:
        activos = activos.filter(ubicacion_id=ubicacion)
    
    # Datos para filtros
    estados = Estado.objects.all()
    categorias = Categoria.objects.all()
    ubicaciones = Ubicacion.objects.all()
    
    context = {
        'activos': activos,
        'estados': estados,
        'categorias': categorias,
        'ubicaciones': ubicaciones,
        'filtro_estado': estado,
        'filtro_categoria': categoria,
        'filtro_ubicacion': ubicacion,
    }
    return render(request, 'reportes/inventario.html', context)

# ============= EXPORTACIÓN A EXCEL =============
@login_required
def exportar_inventario_excel(request):
    """Exportar inventario a Excel"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Código', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Estado', 
               'Ubicación', 'Responsable', 'Fecha Compra', 'Valor', 'Garantía']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Datos
    activos = Activo.objects.select_related(
        'tipo', 'categoria', 'marca', 'estado', 'ubicacion', 'responsable'
    ).all()
    
    for row, activo in enumerate(activos, 2):
        ws.cell(row=row, column=1, value=activo.codigo)
        ws.cell(row=row, column=2, value=str(activo.tipo))
        ws.cell(row=row, column=3, value=activo.marca.nombre)
        ws.cell(row=row, column=4, value=activo.modelo)
        ws.cell(row=row, column=5, value=activo.serial or '')
        ws.cell(row=row, column=6, value=activo.estado.nombre)
        ws.cell(row=row, column=7, value=activo.ubicacion.nombre)
        ws.cell(row=row, column=8, value=str(activo.responsable) if activo.responsable else '')
        ws.cell(row=row, column=9, value=activo.fecha_compra.strftime('%d/%m/%Y') if activo.fecha_compra else '')
        ws.cell(row=row, column=10, value=float(activo.valor_compra) if activo.valor_compra else 0)
        ws.cell(row=row, column=11, value=f"{activo.garantia_meses} meses")
    
    # Ajustar ancho de columnas
    for col in range(1, 12):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    
    wb.save(response)
    return response

# ============= EXPORTACIÓN A PDF =============
@login_required
def exportar_inventario_pdf(request):
    """Exportar inventario a PDF"""
    
    activos = Activo.objects.select_related(
        'tipo', 'categoria', 'marca', 'estado', 'ubicacion', 'responsable'
    ).all()
    
    # Estadísticas
    total_activos = activos.count()
    valor_total = sum(a.valor_compra or 0 for a in activos)
    
    template = get_template('reportes/inventario_pdf.html')
    html = template.render({
        'activos': activos,
        'total_activos': total_activos,
        'valor_total': valor_total,
        'fecha': timezone.now(),
        'usuario': request.user
    })
    
    # Crear PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
        return response
    
    return HttpResponse('Error al generar PDF', status=500)

# ============= REPORTE DE ASIGNACIONES =============
@login_required
def reporte_asignaciones(request):
    """Reporte de asignaciones activas e históricas"""
    
    asignaciones = Asignacion.objects.select_related(
        'activo', 'usuario_asignado', 'asignado_por'
    ).all()
    
    # Filtros
    estado = request.GET.get('estado', 'todas')
    if estado == 'activas':
        asignaciones = asignaciones.filter(activo_actual=True)
    elif estado == 'devueltas':
        asignaciones = asignaciones.filter(activo_devuelto=True)
    
    context = {
        'asignaciones': asignaciones,
        'estado_filtro': estado,
    }
    return render(request, 'reportes/asignaciones.html', context)

@login_required
def exportar_asignaciones_excel(request):
    """Exportar asignaciones a Excel"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asignaciones"
    
    # Headers
    headers = ['ID', 'Activo', 'Asignado a', 'Asignado por', 'Fecha Asignación', 
               'Fecha Devolución', 'Estado', 'Motivo']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    asignaciones = Asignacion.objects.select_related(
        'activo', 'usuario_asignado', 'asignado_por'
    ).all()
    
    for row, asig in enumerate(asignaciones, 2):
        ws.cell(row=row, column=1, value=asig.id)
        ws.cell(row=row, column=2, value=asig.activo.codigo)
        ws.cell(row=row, column=3, value=str(asig.usuario_asignado))
        ws.cell(row=row, column=4, value=str(asig.asignado_por))
        ws.cell(row=row, column=5, value=asig.fecha_asignacion.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=6, value=asig.fecha_devolucion.strftime('%d/%m/%Y %H:%M') if asig.fecha_devolucion else '')
        ws.cell(row=row, column=7, value='Activa' if asig.activo_actual else 'Devuelta')
        ws.cell(row=row, column=8, value=asig.motivo or '')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=asignaciones_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

# ============= GRÁFICOS (DATOS PARA AJAX) =============
@login_required
def datos_grafico_estados(request):
    """Datos para gráfico de activos por estado"""
    from django.http import JsonResponse
    
    estados = Estado.objects.all()
    data = {
        'labels': [],
        'values': [],
        'colors': []
    }
    
    for estado in estados:
        count = Activo.objects.filter(estado=estado).count()
        if count > 0:
            data['labels'].append(estado.nombre)
            data['values'].append(count)
            data['colors'].append(estado.color)
    
    return JsonResponse(data)

@login_required
def datos_grafico_categorias(request):
    """Datos para gráfico de activos por categoría"""
    from django.http import JsonResponse
    
    categorias = Categoria.objects.annotate(
        total=Count('activo')
    ).filter(total__gt=0).order_by('-total')[:10]
    
    data = {
        'labels': [c.nombre for c in categorias],
        'values': [c.total for c in categorias],
    }
    return JsonResponse(data)

@login_required
def datos_grafico_mensual(request):
    """Datos para gráfico de movimientos por mes"""
    from django.http import JsonResponse
    from django.db.models import Count
    from django.db.models.functions import TruncMonth
    
    movimientos = Movimiento.objects.annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        total=Count('id')
    ).order_by('mes')[:12]
    
    data = {
        'labels': [m['mes'].strftime('%b %Y') for m in movimientos if m['mes']],
        'values': [m['total'] for m in movimientos],
    }
    return JsonResponse(data)

@login_required
def datos_grafico_tipos(request):
    """Datos para gráfico de activos por tipo"""
    from django.http import JsonResponse
    from catalogos.models import TipoActivo
    
    tipos = TipoActivo.objects.annotate(
        total=Count('activo')
    ).filter(total__gt=0).order_by('-total')
    
    # Paleta de colores profesional
    colores = [
        '#1A73E8', '#4285F4', '#5B9CF5', '#85C1E2', '#ADD8E6',  # Azules
        '#DC3912', '#E8524D', '#F4744E', '#FF6B6B', '#FF8C8C',  # Rojos
        '#FF9800', '#FFB74D', '#FFCC80', '#FFD54F', '#FFEB3B',  # Naranjas/Amarillos
        '#66BB6A', '#81C784', '#A5D6A7', '#C8E6C9', '#E8F5E9',  # Verdes
        '#AB47BC', '#BA68C8', '#CE93D8', '#E1BEE7', '#F3E5F5',  # Púrpuras
        '#29B6F6', '#42A5F5', '#64B5F6', '#90CAF9', '#BBDEFB',  # Azul claro
        '#7E57C2', '#9575CD', '#B39DDB', '#D1C4E9', '#EDE7F6',  # Púrpura
    ]
    
    data = {
        'labels': [t.nombre for t in tipos],
        'values': [t.total for t in tipos],
        'colors': [colores[i % len(colores)] for i in range(len(tipos))]
    }
    return JsonResponse(data)

# ============= REPORTE DE CATEGORÍAS =============
@login_required
def reporte_categorias(request):
    """Reporte de activos agrupados por categoría"""
    
    categorias = Categoria.objects.annotate(
        total_activos=Count('activo'),
        activos_asignados=Count('activo', filter=Q(activo__estado__nombre='ASIGNADO')),
        activos_disponibles=Count('activo', filter=Q(activo__estado__nombre='DISPONIBLE')),
        valor_total=Sum('activo__valor_compra')
    ).order_by('-total_activos')
    
    context = {
        'categorias': categorias,
        'total_general': sum(c.total_activos for c in categorias),
        'valor_general': sum(c.valor_total or 0 for c in categorias),
    }
    return render(request, 'reportes/categorias.html', context)

@login_required
def exportar_categorias_excel(request):
    """Exportar reporte de categorías a Excel"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    
    # Headers
    headers = ['Categoría', 'Total Activos', 'Asignados', 'Disponibles', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    categorias = Categoria.objects.annotate(
        total_activos=Count('activo'),
        activos_asignados=Count('activo', filter=Q(activo__estado__nombre='ASIGNADO')),
        activos_disponibles=Count('activo', filter=Q(activo__estado__nombre='DISPONIBLE')),
        valor_total=Sum('activo__valor_compra')
    )
    
    for row, cat in enumerate(categorias, 2):
        ws.cell(row=row, column=1, value=cat.nombre)
        ws.cell(row=row, column=2, value=cat.total_activos)
        ws.cell(row=row, column=3, value=cat.activos_asignados)
        ws.cell(row=row, column=4, value=cat.activos_disponibles)
        ws.cell(row=row, column=5, value=float(cat.valor_total or 0))
    
    # Ajustar ancho
    for col in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=categorias_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

# ============= REPORTE DE MOVIMIENTOS =============
@login_required
def reporte_movimientos(request):
    """Reporte de movimientos con filtros"""
    
    movimientos = Movimiento.objects.select_related(
        'activo', 'usuario'
    ).all().order_by('-fecha')
    
    # Filtros
    tipo = request.GET.get('tipo')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
    if fecha_desde:
        movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)
    
    context = {
        'movimientos': movimientos,
        'tipos': Movimiento.TIPO_MOVIMIENTO,
    }
    return render(request, 'reportes/movimientos.html', context)

@login_required
def exportar_movimientos_excel(request):
    """Exportar movimientos a Excel"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    headers = ['Fecha', 'Activo', 'Tipo', 'Usuario', 'Origen', 'Destino', 'Observaciones']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    movimientos = Movimiento.objects.select_related(
        'activo', 'usuario'
    ).all().order_by('-fecha')
    
    for row, mov in enumerate(movimientos, 2):
        ws.cell(row=row, column=1, value=mov.fecha.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=mov.activo.codigo)
        ws.cell(row=row, column=3, value=mov.get_tipo_display())
        ws.cell(row=row, column=4, value=str(mov.usuario))
        ws.cell(row=row, column=5, value=mov.ubicacion_origen.nombre if mov.ubicacion_origen else '')
        ws.cell(row=row, column=6, value=mov.ubicacion_destino.nombre if mov.ubicacion_destino else '')
        ws.cell(row=row, column=7, value=mov.observaciones or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=movimientos_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

# ============= REPORTE POR RESPONSABLE =============
@login_required
def reporte_por_responsable(request, user_id):
    """Reporte de activos asignados a un responsable específico"""
    
    responsable = get_object_or_404(User, pk=user_id)
    activos = Activo.objects.filter(
        responsable=responsable
    ).select_related('tipo', 'categoria', 'marca', 'estado', 'ubicacion')
    
    context = {
        'responsable': responsable,
        'activos': activos,
        'total': activos.count(),
        'valor_total': sum(a.valor_compra or 0 for a in activos),
    }
    return render(request, 'reportes/por_responsable.html', context)

@login_required
def exportar_responsable_excel(request, user_id):
    """Exportar activos de un responsable a Excel"""
    
    responsable = get_object_or_404(User, pk=user_id)
    activos = Activo.objects.filter(responsable=responsable)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Activos de {responsable.username}"
    
    headers = ['Código', 'Tipo', 'Marca', 'Modelo', 'Serial', 'Estado', 'Ubicación']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row, activo in enumerate(activos, 2):
        ws.cell(row=row, column=1, value=activo.codigo)
        ws.cell(row=row, column=2, value=str(activo.tipo))
        ws.cell(row=row, column=3, value=activo.marca.nombre)
        ws.cell(row=row, column=4, value=activo.modelo)
        ws.cell(row=row, column=5, value=activo.serial or '')
        ws.cell(row=row, column=6, value=activo.estado.nombre)
        ws.cell(row=row, column=7, value=activo.ubicacion.nombre)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=responsable_{responsable.username}_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_asignaciones_pdf(request):
    """Exportar asignaciones a PDF"""
    
    asignaciones = Asignacion.objects.select_related(
        'activo', 'usuario_asignado', 'asignado_por'
    ).all()
    
    # Estadísticas
    total_asignaciones = asignaciones.count()
    activas = asignaciones.filter(activo_actual=True).count()
    devueltas = asignaciones.filter(activo_devuelto=True).count()
    
    template = get_template('reportes/asignaciones_pdf.html')
    html = template.render({
        'asignaciones': asignaciones,
        'total_asignaciones': total_asignaciones,
        'activas': activas,
        'devueltas': devueltas,
        'fecha': timezone.now(),
        'usuario': request.user
    })
    
    # Crear PDF
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=asignaciones_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
        return response
    
    return HttpResponse('Error al generar PDF', status=500)